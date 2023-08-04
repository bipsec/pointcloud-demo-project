import os
import random
import openpyxl
import tkinter as tk

from tkinter import filedialog

from pointcloud import PointCloudApp


class PointCloudHandler:
    def __init__(self, num_points=100):
        self.num_points = num_points
        self.point_cloud_data = self.generate_synthetic_point_cloud()

    def generate_synthetic_point_cloud(self):
        point_cloud = []
        for _ in range(self.num_points):
            x = random.uniform(-10, 10)
            y = random.uniform(-10, 10)
            z = random.uniform(-10, 10)
            point_cloud.append((x, y, z))
        return point_cloud

    def save_to_excel(self, sheet_name="PointCloudSheet"):
        excel_file_path = "synthetic_point_cloud.xlsx"

        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet(title=sheet_name)

        for index, (x, y, z) in enumerate(self.point_cloud_data, start=1):
            sheet.cell(row=index, column=1, value=x)
            sheet.cell(row=index, column=2, value=y)
            sheet.cell(row=index, column=3, value=z)

        workbook.save(excel_file_path)
        print(f"Data saved to {excel_file_path} successfully.")

    def import_from_excel(self, sheet_name="PointCloudSheet"):
        excel_file_path = filedialog.askopenfilename(initialdir=".", title="Select Excel File",
                                                     filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))

        if not excel_file_path:
            print("No file selected.")
            return

        if not os.path.exists(excel_file_path):
            print("Error: Excel file not found.")
            return

        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[sheet_name]

        self.point_cloud_data = []
        for row in sheet.iter_rows(values_only=True):
            if len(row) >= 3:
                x, y, z = row[0], row[1], row[2]
                self.point_cloud_data.append((x, y, z))

        print("Data imported from Excel successfully.")


if __name__ == "__main__":
    num_points = 100
    point_cloud_handler = PointCloudHandler(num_points=num_points)

    root = tk.Tk()
    app = PointCloudApp(root, point_cloud_handler)
    root.mainloop()
